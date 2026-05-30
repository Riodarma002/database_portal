import streamlit as st
import pandas as pd
import database
from auth import check_password, logout
import time
import os
from streamlit_option_menu import option_menu

# Direktori tempat app.py berada (digunakan untuk path relatif)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

st.set_page_config(page_title="MGE Portal", page_icon="🏢", layout="wide")

# =========================================================================
# GLOBAL CSS UNTUK SIDEBAR & UI
# =========================================================================
st.markdown("""
<style>
    /* Mengurangi ruang kosong (whitespace) di bagian paling atas aplikasi */
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 1rem !important;
    }

    /* Mempercantik dan mengecilkan ukuran header (judul) */
    h1 {
        color: #2E7D32 !important;
        font-weight: 800 !important;
        font-size: 2.0rem !important; /* Ukuran diperkecil */
        padding-top: 0rem !important;
        margin-top: -1rem !important; /* Ditarik ke atas */
        margin-bottom: 0.5rem !important;
    }
    
    /* Memperkecil elemen Filter (Input, Select, Button) */
    div[data-testid="stExpander"] input, 
    div[data-testid="stExpander"] div[data-baseweb="select"] > div {
        font-size: 13px !important;
        min-height: 32px !important;
        padding-top: 0px !important;
        padding-bottom: 0px !important;
    }
    
    div[data-testid="stExpander"] button {
        padding: 2px 15px !important;
        font-size: 14px !important;
        min-height: 36px !important;
        border-radius: 6px !important;
    }
    
    div[data-testid="stExpander"] label {
        font-size: 13px !important;
        padding-bottom: 2px !important;
    }
    
    /* Modern Tab Navigation Styling (Top-level tabs) */
    div[data-baseweb="tab-list"] {
        gap: 8px;
    }
    button[data-baseweb="tab"] {
        font-size: 15px !important;
        font-weight: 600 !important;
        padding: 12px 24px !important;
        border-radius: 6px 6px 0 0 !important;
        transition: all 0.2s ease-in-out !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        background-color: #1F4E78 !important;
        color: white !important;
    }
    button[data-baseweb="tab"][aria-selected="false"]:hover {
        background-color: #f0f2f6 !important;
        color: #1F4E78 !important;
    }
    
    /* Professional Sidebar Logout Button */
    div[data-testid="stSidebar"] div.stButton > button {
        border: 1.5px solid #d32f2f !important;
        color: #d32f2f !important;
        background-color: transparent !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        transition: all 0.2s ease-in-out !important;
    }
    div[data-testid="stSidebar"] div.stButton > button:hover {
        background-color: #d32f2f !important;
        color: white !important;
        border: 1.5px solid #d32f2f !important;
    }
</style>
""", unsafe_allow_html=True)

# Cek Login
if not check_password():
    st.stop()

# =========================================================================
# SIDEBAR MENU PUSAT
# =========================================================================
# Daftar Modul Tanpa Emoji (Digantikan oleh ikon Bootstrap di option_menu)
module_list = [
    "Dashboard Utama", 
    "Modul Coal Hauling", 
    "Modul Overburden (OB)", 
    "Modul MS Kontrak", 
    "Modul Fuel"
]

# Baca memori posisi menu dari URL (hanya saat pertama kali load)
if "active_menu_idx" not in st.session_state:
    active_module = st.query_params.get("module", "0")
    try:
        st.session_state["active_menu_idx"] = int(active_module)
        if st.session_state["active_menu_idx"] >= len(module_list):
            st.session_state["active_menu_idx"] = 0
    except:
        st.session_state["active_menu_idx"] = 0

with st.sidebar:
    import base64
    try:
        logo_path = os.path.join(BASE_DIR, ".streamlit", "logo_mge.png")
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f"""
            <div style="margin-bottom: 15px; display: flex; justify-content: center;">
                <img src="data:image/png;base64,{logo_b64}" style="height: 45px; object-fit: contain;">
            </div>
        """, unsafe_allow_html=True)
    except:
        pass
    st.markdown("<div style='margin-top: 15px;'>Selamat datang, <b>Admin Planning</b></div>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    
    menu = option_menu(
        menu_title=None,
        options=module_list,
        icons=["house-door", "truck", "cone-striped", "file-earmark-text", "fuel-pump"],
        menu_icon="cast",
        default_index=st.session_state["active_menu_idx"],
        key="sidebar_menu",
        styles={
            "container": {"padding": "0!important", "background-color": "transparent", "border": "none"},
            "icon": {"color": "#666", "font-size": "18px"}, 
            "nav-link": {
                "font-size": "14px", 
                "text-align": "left", 
                "margin":"0px", 
                "padding": "12px",
                "color": "#444",
                "font-weight": "500",
                "font-family": "'Plus Jakarta Sans', sans-serif"
            },
            "nav-link-selected": {
                "background-color": "#e0eaf5", 
                "color": "#1F4E78", 
                "font-weight": "700",
                "border-left": "4px solid #1F4E78",
                "border-radius": "0px"
            },
        }
    )
    
    # Simpan posisi menu ke session_state & query_params (tanpa rerun)
    try:
        current_idx = module_list.index(menu)
        st.session_state["active_menu_idx"] = current_idx
        st.query_params["module"] = str(current_idx)
    except:
        pass
        
    st.markdown("---")
    st.button(":material/logout: Logout", on_click=logout, use_container_width=True)


@st.cache_data(ttl=600)
def get_unique_vendors():
    df = database.fetch_data("SELECT DISTINCT vendor FROM coal_hauling WHERE vendor IS NOT NULL AND vendor != '' ORDER BY vendor")
    if not df.empty:
        return ["Semua"] + df['vendor'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_unit_types():
    df = database.fetch_data("SELECT DISTINCT unit_type FROM coal_hauling WHERE unit_type IS NOT NULL AND unit_type != '' ORDER BY unit_type")
    if not df.empty:
        return ["Semua"] + df['unit_type'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_pits():
    df = database.fetch_data("SELECT DISTINCT pit FROM coal_hauling WHERE pit IS NOT NULL AND pit != '' ORDER BY pit")
    if not df.empty:
        return ["Semua"] + df['pit'].tolist()
    return ["Semua"]



@st.cache_data(ttl=600)
def get_unique_transit_pits():
    df = database.fetch_data("SELECT DISTINCT pit FROM coal_transit WHERE pit IS NOT NULL AND pit != '' ORDER BY pit")
    if not df.empty:
        return ["Semua"] + df['pit'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_transit_diggers():
    df = database.fetch_data("SELECT DISTINCT digger FROM coal_transit WHERE digger IS NOT NULL AND digger != '' ORDER BY digger")
    if not df.empty:
        return ["Semua"] + df['digger'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_transit_unit_codes():
    df = database.fetch_data("SELECT DISTINCT unit_code FROM coal_transit WHERE unit_code IS NOT NULL AND unit_code != '' ORDER BY unit_code")
    if not df.empty:
        return ["Semua"] + df['unit_code'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def fetch_transit_data(date_start=None, date_end=None):
    query = """
    SELECT id, date, shift, unit_code, model, type, brand, user, seam, block, product_code, product_inv, pit, digger, `1`, `2`, `3`, `4`, `5`, `6`, `7`, `8`, `9`, `10`, `11`, `12`, total, vessel, netto, periode, room 
    FROM coal_transit 
    WHERE 1=1
    """
    params = []
    
    if date_start and date_end:
        query += " AND date BETWEEN %s AND %s"
        params.extend([date_start, date_end])
    elif date_start:
        query += " AND date = %s"
        params.append(date_start)
    
    if date_start:
        query += " ORDER BY date DESC, id DESC"
    else:
        query += " ORDER BY date DESC, id DESC LIMIT 1000"
    
    return database.fetch_data(query, tuple(params) if params else None)


@st.cache_data(ttl=600)
def fetch_hauling_data(date_start=None, date_end=None):
    """Mengambil data hauling dari database. Di-cache berdasarkan rentang tanggal.
    Filter lainnya (shift, pit, vendor, dll) diterapkan client-side via Pandas."""
    query = """
    SELECT jml, day, date, shift, loading_date, voucher_number, pit, block, seam, product, concat, vendor, unit_type, unit_id, payload_arrival_time, payload_embark_time, weight_gross, weight_empty, weight_nett, destination, route, loader_id, tonage 
    FROM coal_hauling 
    WHERE 1=1
    """
    params = []
    
    if date_start and date_end:
        query += " AND date BETWEEN %s AND %s"
        params.extend([date_start, date_end])
    elif date_start:
        query += " AND date = %s"
        params.append(date_start)
    
    if date_start:
        query += " ORDER BY date DESC, id DESC"
    else:
        query += " ORDER BY date DESC, id DESC LIMIT 1000"
    
    return database.fetch_data(query, tuple(params) if params else None)

def clear_all_cache():
    """Menghapus semua cache data setelah operasi upload/delete."""
    fetch_hauling_data.clear()
    get_unique_vendors.clear()
    get_unique_unit_types.clear()
    get_unique_pits.clear()
    
    fetch_fuel_data.clear()
    get_unique_vendors_fuel.clear()
    get_unique_units_fuel.clear()
    get_unique_locations_fuel.clear()
    
    get_ob_unique_pits.clear()
    get_ob_unique_owners.clear()
    get_ob_unique_shifts.clear()
    
    get_dashboard_summary.clear()

@st.cache_data(ttl=600)
def get_ob_unique_pits():
    df = database.fetch_data("SELECT DISTINCT pit FROM ob_ob WHERE pit IS NOT NULL AND pit != '' ORDER BY pit")
    if not df.empty:
        return ["Semua"] + df['pit'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_ob_unique_shifts():
    df = database.fetch_data("SELECT DISTINCT shift FROM ob_ob WHERE shift IS NOT NULL AND shift != '' ORDER BY shift")
    if not df.empty:
        return ["Semua"] + df['shift'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_ob_unique_owners():
    df = database.fetch_data("SELECT DISTINCT owner FROM ob_ob WHERE owner IS NOT NULL AND owner != '' ORDER BY owner")
    if not df.empty:
        return ["Semua"] + df['owner'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_dashboard_summary():
    last_hauling_str = "Belum Tersedia"
    last_fuel_str = "Belum Tersedia"
    last_ob_str = "Belum Tersedia"
    last_ms_str = "Belum Tersedia"
    hauling_yearly = pd.DataFrame()
    fuel_yearly = pd.DataFrame()
    hauling_daily = pd.DataFrame()
    fuel_daily = pd.DataFrame()
    ob_yearly = pd.DataFrame()
    ob_daily = pd.DataFrame()

    try:
        df_hauling = database.fetch_data("SELECT MAX(date) as last_date FROM coal_hauling")
        if not df_hauling.empty and pd.notnull(df_hauling.iloc[0]['last_date']):
            last_hauling_str = df_hauling.iloc[0]['last_date'].strftime('%d %b %Y')
    except:
        pass

    try:
        df_fuel = database.fetch_data("SELECT MAX(date) as last_date FROM fuel")
        if not df_fuel.empty and pd.notnull(df_fuel.iloc[0]['last_date']):
            last_fuel_str = df_fuel.iloc[0]['last_date'].strftime('%d %b %Y')
    except:
        pass

    try:
        df_ob = database.fetch_data("SELECT MAX(date) as last_date FROM ob_ob")
        if not df_ob.empty and pd.notnull(df_ob.iloc[0]['last_date']):
            last_ob_str = df_ob.iloc[0]['last_date'].strftime('%d %b %Y')
    except:
        pass

    try:
        hauling_yearly = database.fetch_data("""
            SELECT DATE_FORMAT(date, '%Y-%m') as month, SUM(tonage) as total 
            FROM coal_hauling 
            GROUP BY month 
            ORDER BY month DESC LIMIT 12
        """)
        if not hauling_yearly.empty:
            hauling_yearly = hauling_yearly.sort_values('month')
    except:
        pass

    try:
        fuel_yearly = database.fetch_data("""
            SELECT DATE_FORMAT(date, '%Y-%m') as month, SUM(refueling) as total 
            FROM fuel 
            GROUP BY month 
            ORDER BY month DESC LIMIT 12
        """)
        if not fuel_yearly.empty:
            fuel_yearly = fuel_yearly.sort_values('month')
    except:
        pass

    try:
        hauling_daily = database.fetch_data("""
            SELECT date, SUM(tonage) as total 
            FROM coal_hauling 
            WHERE date >= (SELECT DATE_SUB(MAX(date), INTERVAL 30 DAY) FROM coal_hauling)
            GROUP BY date 
            ORDER BY date ASC
        """)
        if not hauling_daily.empty:
            hauling_daily['date'] = pd.to_datetime(hauling_daily['date']).dt.strftime('%d %b')
    except:
        pass

    try:
        fuel_daily = database.fetch_data("""
            SELECT date, SUM(refueling) as total 
            FROM fuel 
            WHERE date >= (SELECT DATE_SUB(MAX(date), INTERVAL 30 DAY) FROM fuel)
            GROUP BY date 
            ORDER BY date ASC
        """)
        if not fuel_daily.empty:
            fuel_daily['date'] = pd.to_datetime(fuel_daily['date']).dt.strftime('%d %b')
    except:
        pass

    try:
        ob_yearly = database.fetch_data("""
            SELECT DATE_FORMAT(date, '%Y-%m') as month, SUM(volume_js) as total 
            FROM ob_volume_ob_by_js 
            GROUP BY month 
            ORDER BY month DESC LIMIT 12
        """)
        if not ob_yearly.empty:
            ob_yearly = ob_yearly.sort_values('month')
    except:
        pass

    try:
        ob_daily = database.fetch_data("""
            SELECT date, SUM(volume_js) as total 
            FROM ob_volume_ob_by_js 
            WHERE date >= (SELECT DATE_SUB(MAX(date), INTERVAL 30 DAY) FROM ob_volume_ob_by_js)
            GROUP BY date 
            ORDER BY date ASC
        """)
        if not ob_daily.empty:
            ob_daily['date'] = pd.to_datetime(ob_daily['date']).dt.strftime('%d %b')
    except:
        pass

    return last_hauling_str, last_fuel_str, last_ob_str, last_ms_str, hauling_yearly, fuel_yearly, hauling_daily, fuel_daily, ob_yearly, ob_daily


@st.cache_data(ttl=600)
def get_unique_vendors_fuel():
    df = database.fetch_data("SELECT DISTINCT vendor FROM fuel WHERE vendor IS NOT NULL AND vendor != '' ORDER BY vendor")
    if not df.empty:
        return ["Semua"] + df['vendor'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_units_fuel():
    df = database.fetch_data("SELECT DISTINCT unit_code FROM fuel WHERE unit_code IS NOT NULL AND unit_code != '' ORDER BY unit_code")
    if not df.empty:
        return ["Semua"] + df['unit_code'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def get_unique_locations_fuel():
    df = database.fetch_data("SELECT DISTINCT alocation FROM fuel WHERE alocation IS NOT NULL AND alocation != '' ORDER BY alocation")
    if not df.empty:
        return ["Semua"] + df['alocation'].tolist()
    return ["Semua"]

@st.cache_data(ttl=600)
def fetch_fuel_data(date_start=None, date_end=None):
    """Mengambil data fuel dari database. Di-cache berdasarkan rentang tanggal."""
    query = """
    SELECT id, unit_fix, date, periode, shift, time, reg_no, unit_code, unit_model, unit_type, brand, vendor, alocation, km, hm, fm_awal, fm_akhir, refueling, source, location, operator, fuelman, no_voucher
    FROM fuel 
    WHERE 1=1
    """
    params = []
    
    if date_start and date_end:
        query += " AND date BETWEEN %s AND %s"
        params.extend([date_start, date_end])
    elif date_start:
        query += " AND date = %s"
        params.append(date_start)
    
    if date_start:
        query += " ORDER BY date DESC, id DESC"
    else:
        query += " ORDER BY date DESC, id DESC LIMIT 1000"
    
    return database.fetch_data(query, tuple(params) if params else None)

# =========================================================================
# KONTEN BERDASARKAN MENU YANG DIPILIH
# =========================================================================

if menu == "Dashboard Utama":
    st.title("Dashboard Utama MGE")
    st.markdown("Ringkasan seluruh aktivitas operasional (Hauling, OB, MS, Fuel).")
    
    last_hauling_str, last_fuel_str, last_ob_str, last_ms_str, hauling_yearly, fuel_yearly, hauling_daily, fuel_daily, ob_yearly, ob_daily = get_dashboard_summary()
    
    # Modern Metric Cards HTML
    st.markdown(f"""
    <style>
    .metric-card {{
        background-color: #ffffff;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 4px 10px rgba(0,0,0,0.05);
        border: 1px solid #f0f0f0;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }}
    .metric-card:hover {{
        transform: translateY(-3px);
        box-shadow: 0 8px 15px rgba(0,0,0,0.1);
    }}
    .metric-icon {{
        font-size: 28px;
        margin-right: 15px;
        width: 60px;
        height: 60px;
        display: flex;
        align-items: center;
        justify-content: center;
        border-radius: 12px;
    }}
    .icon-hauling {{ background-color: #e0f2fe; color: #0284c7; }}
    .icon-fuel {{ background-color: #fee2e2; color: #dc2626; }}
    .icon-ob {{ background-color: #fef3c7; color: #d97706; }}
    .icon-ms {{ background-color: #dcfce7; color: #16a34a; }}
    
    .metric-content {{ flex-grow: 1; }}
    .metric-title {{
        font-size: 13px;
        color: #64748b;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 4px;
        font-family: 'Plus Jakarta Sans', sans-serif;
    }}
    .metric-value {{
        font-size: 20px;
        font-weight: 700;
        color: #0f172a;
        font-family: 'Plus Jakarta Sans', sans-serif;
    }}
    </style>
    
    <div style="display: flex; gap: 15px; flex-wrap: wrap;">
        <div style="flex: 1; min-width: 200px;" class="metric-card">
            <div class="metric-icon icon-hauling">🚛</div>
            <div class="metric-content">
                <div class="metric-title">Last Sync Hauling</div>
                <div class="metric-value">{last_hauling_str}</div>
            </div>
        </div>
        <div style="flex: 1; min-width: 200px;" class="metric-card">
            <div class="metric-icon icon-fuel">⛽</div>
            <div class="metric-content">
                <div class="metric-title">Last Sync Fuel</div>
                <div class="metric-value">{last_fuel_str}</div>
            </div>
        </div>
        <div style="flex: 1; min-width: 200px;" class="metric-card">
            <div class="metric-icon icon-ob">🚧</div>
            <div class="metric-content">
                <div class="metric-title">Last Sync OB</div>
                <div class="metric-value">{last_ob_str}</div>
            </div>
        </div>
        <div style="flex: 1; min-width: 200px;" class="metric-card">
            <div class="metric-icon icon-ms">📝</div>
            <div class="metric-content">
                <div class="metric-title">Last Sync MS</div>
                <div class="metric-value">{last_ms_str}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    tab1, tab2, tab3, tab4 = st.tabs(["🚛 Coal Hauling", "⛽ Fuel", "🚧 Overburden (OB)", "📝 MS Contract"])
    
    import plotly.express as px
    
    with tab1:
        st.subheader("Performance Coal Hauling")
        c1, c2 = st.columns(2)
        with c1:
            if not hauling_yearly.empty:
                fig1 = px.bar(hauling_yearly, x='month', y='total', title='Trend Tonase 12 Bulan Terakhir',
                              labels={'month': 'Bulan', 'total': 'Total Tonase (Ton)'},
                              text='total',
                              color_discrete_sequence=['#2563eb'])
                fig1.update_traces(
                    texttemplate='%{text:.2s}', 
                    textposition='outside', 
                    marker_line_width=0,
                    textfont=dict(size=11, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Tonase: %{y:,.0f} Ton<extra></extra>"
                )
                fig1.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", dtick="M1", tickformat="%b<br>%Y"),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig1, use_container_width=True)
            else:
                st.info("Data bulanan Coal Hauling belum tersedia.")
        
        with c2:
            if not hauling_daily.empty:
                fig2 = px.area(hauling_daily, x='date', y='total', title='Tonase Harian (30 Hari Terakhir)',
                               labels={'date': 'Tanggal', 'total': 'Tonase Harian (Ton)'},
                               color_discrete_sequence=['#10b981'])
                fig2.update_traces(
                    mode='lines+markers+text', 
                    text=hauling_daily['total'], 
                    texttemplate='%{text:.2s}', 
                    textposition='top center',
                    line_shape='spline',
                    fillcolor='rgba(16, 185, 129, 0.15)',
                    line=dict(width=3),
                    marker=dict(size=6, color='#10b981', line=dict(width=2, color='white')),
                    textfont=dict(size=10, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Tonase: %{y:,.0f} Ton<extra></extra>"
                )
                fig2.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", tickangle=-45),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("Data harian Coal Hauling belum tersedia.")
                
    with tab2:
        st.subheader("Performance Fuel")
        c1, c2 = st.columns(2)
        with c1:
            if not fuel_yearly.empty:
                fig3 = px.line(fuel_yearly, x='month', y='total', title='Trend Konsumsi Fuel 12 Bulan Terakhir',
                               text='total',
                               labels={'month': 'Bulan', 'total': 'Total Fuel (Liter)'},
                               color_discrete_sequence=['#f43f5e'])
                fig3.update_traces(
                    mode='lines+markers+text',
                    texttemplate='%{text:.2s}', 
                    textposition='top center',
                    line_shape='spline',
                    line=dict(width=3),
                    marker=dict(size=7, color='#f43f5e', line=dict(width=2, color='white')),
                    textfont=dict(size=11, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Fuel: %{y:,.0f} Liter<extra></extra>"
                )
                fig3.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", dtick="M1", tickformat="%b<br>%Y"),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("Data bulanan Fuel belum tersedia.")
        
        with c2:
            if not fuel_daily.empty:
                fig4 = px.bar(fuel_daily, x='date', y='total', title='Konsumsi Fuel Harian (30 Hari Terakhir)',
                               text='total',
                               labels={'date': 'Tanggal', 'total': 'Konsumsi Harian (Liter)'},
                               color_discrete_sequence=['#f59e0b'])
                fig4.update_traces(
                    texttemplate='%{text:.2s}', 
                    textposition='outside', 
                    marker_line_width=0,
                    marker_color='rgba(245, 158, 11, 0.85)',
                    textfont=dict(size=10, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Fuel: %{y:,.0f} Liter<extra></extra>"
                )
                fig4.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", tickangle=-45),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info("Data harian Fuel belum tersedia.")
                
    with tab3:
        st.subheader("Performance Overburden (OB)")
        c1, c2 = st.columns(2)
        with c1:
            if not ob_yearly.empty:
                fig5 = px.line(ob_yearly, x='month', y='total', title='Trend Volume OB (BCM) 12 Bulan Terakhir',
                               text='total',
                               labels={'month': 'Bulan', 'total': 'Volume (BCM)'},
                               color_discrete_sequence=['#d97706'])
                fig5.update_traces(
                    mode='lines+markers+text',
                    texttemplate='%{text:.3s}', 
                    textposition='top center',
                    line_shape='spline',
                    line=dict(width=3),
                    marker=dict(size=7, color='#d97706', line=dict(width=2, color='white')),
                    textfont=dict(size=11, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Volume: %{y:,.0f} BCM<extra></extra>"
                )
                fig5.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", dtick="M1", tickformat="%b<br>%Y"),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig5, use_container_width=True)
            else:
                st.info("Data bulanan Overburden belum tersedia.")
                
        with c2:
            if not ob_daily.empty:
                fig6 = px.bar(ob_daily, x='date', y='total', title='Volume OB Harian (30 Hari Terakhir)',
                               text='total',
                               labels={'date': 'Tanggal', 'total': 'Volume (BCM)'},
                               color_discrete_sequence=['#fbbf24'])
                fig6.update_traces(
                    texttemplate='%{text:.3s}', 
                    textposition='outside', 
                    marker_line_width=0,
                    marker_color='rgba(251, 191, 36, 0.85)',
                    textfont=dict(size=10, color="#64748b"),
                    hovertemplate="<b>%{x}</b><br>Volume: %{y:,.0f} BCM<extra></extra>"
                )
                fig6.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                    xaxis=dict(showgrid=False, title="", tickangle=-45),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title=""),
                    font=dict(family="Plus Jakarta Sans, sans-serif", color="#475569"),
                    title_font=dict(size=16, color="#0f172a", family="Plus Jakarta Sans, sans-serif")
                )
                st.plotly_chart(fig6, use_container_width=True)
            else:
                st.info("Data harian Overburden belum tersedia.")
            
    with tab4:
        st.subheader("Performance Mine Survey (MS) Contract")
        st.info("Placeholder untuk grafik Mine Survey. Data akan tampil jika modul ini sudah diisi.")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**(Placeholder) Volume MS 12 Bulan Terakhir**")
            st.line_chart({"Bulan": ["Jan", "Feb", "Mar"], "Volume": [0, 0, 0]}, x="Bulan", y="Volume")
        with c2:
            st.markdown("**(Placeholder) Progress Volume (30 Hari Terakhir)**")
            st.bar_chart({"Tanggal": ["01", "02", "03"], "Volume": [0, 0, 0]}, x="Tanggal", y="Volume")


elif menu == "Modul Coal Hauling":
    st.title("Modul Coal Hauling & Transit")
    st.markdown("Sinkronisasi data produksi batubara (Coal Hauling dan Coal Transit).")
    
    sub_menu = option_menu(
        menu_title=None,
        options=["Coal Hauling", "Coal Transit"],
        icons=["truck", "arrow-left-right"],
        default_index=0,
        orientation="horizontal",
        styles={
            "container": {"padding": "0!important", "margin-bottom": "20px"},
            "nav-link-selected": {"background-color": "#1F4E78"}
        }
    )
    
    if sub_menu == "Coal Hauling":
        tab1, tab2, tab3 = st.tabs([":material/monitoring: View Data", ":material/cloud_upload: Upload / Import", ":material/delete_sweep: Rollback / Delete Batch"])

        with tab1:

            with st.expander(":material/filter_alt: Filter & Cari Data", expanded=True):
                cols = st.columns([1.5, 0.9, 1, 1, 1, 1.2, 0.7, 0.5, 0.8])
                with cols[0]:
                    f_date = st.date_input("Tanggal", value=[])
                with cols[1]:
                    f_shift = st.selectbox("Shift", ["Semua", "DAY", "NIGHT"])
                with cols[2]:
                    f_pit = st.selectbox("Pit", get_unique_pits())
                with cols[3]:
                    f_vendor = st.selectbox("Vendor", get_unique_vendors())
                with cols[4]:
                    f_unit = st.selectbox("Unit", get_unique_unit_types())
                with cols[5]:
                    f_search = st.text_input("Pencarian", placeholder="Voucher/ID...")
                with cols[6]:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    btn_search = st.button(":material/search: Cari", type="primary", use_container_width=True)
                with cols[7]:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    if st.button(":material/refresh:", use_container_width=True, help="Tarik ulang data segar dari database"):
                        clear_all_cache()
                        st.rerun()
                with cols[8]:
                    download_placeholder = st.container()

            # =====================================================================
            # OPTIMASI: SQL hanya dipanggil saat tanggal berubah (cached).
            # Filter Shift/Pit/Vendor/Unit/Search dikerjakan client-side via Pandas.
            # =====================================================================
            date_start = f_date[0] if len(f_date) >= 1 else None
            date_end = f_date[1] if len(f_date) >= 2 else None

            with st.spinner("Mengambil data..."):
                df = fetch_hauling_data(date_start, date_end)

            # Filter client-side (INSTAN, tanpa query SQL ulang)
            if not df.empty:
                if f_shift != "Semua":
                    df = df[df['shift'] == f_shift]
                if f_pit != "Semua":
                    df = df[df['pit'] == f_pit]
                if f_vendor != "Semua":
                    df = df[df['vendor'] == f_vendor]
                if f_unit != "Semua":
                    df = df[df['unit_type'] == f_unit]
                if f_search:
                    mask = df.astype(str).apply(lambda x: x.str.contains(f_search, case=False, na=False)).any(axis=1)
                    df = df[mask]

            if not df.empty:
                # Perbaiki format timedelta dari MySQL agar tampil sebagai HH:MM:SS
                for col in ['payload_arrival_time', 'payload_embark_time']:
                    if col in df.columns:
                        df[col] = df[col].astype(str).apply(lambda x: x.split()[-1] if x != 'NaT' and x != 'None' else '')

                # Pastikan kolom berat adalah angka (float) agar bisa diformat dengan tepat
                for num_col in ['weight_gross', 'weight_empty', 'weight_nett', 'tonage']:
                    if num_col in df.columns:
                        df[num_col] = pd.to_numeric(df[num_col], errors='coerce')
                        # Jika data ter-upload dalam format kg (ribuan), normalisasi kembali ke desimal ton
                        if num_col != 'tonage':
                            df[num_col] = df[num_col].apply(lambda x: x / 1000 if pd.notnull(x) and x >= 1000 else x)

                # Pastikan kolom tanggal berformat datetime agar bisa diatur tampilannya
                for date_col in ['date', 'loading_date']:
                    if date_col in df.columns:
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

                # -------------------------------------------------------------
                # Tombol Download Custom untuk memastikan format Excel sempurna
                # -------------------------------------------------------------
                df_export = df.copy()
                # Kunci 3 angka di belakang koma (tanpa pembulatan aneh dari Pandas)
                for col in ['weight_gross', 'weight_empty', 'weight_nett']:
                    if col in df_export.columns:
                        df_export[col] = df_export[col].apply(lambda x: f"{float(x):.3f}" if pd.notnull(x) else "")
                # Kunci 2 angka di belakang koma untuk tonase
                if 'tonage' in df_export.columns:
                    df_export['tonage'] = df_export['tonage'].apply(lambda x: f"{float(x):.2f}" if pd.notnull(x) else "")

                # Format tanggal juga dikembalikan ke format kalender normal
                for d_col in ['date', 'loading_date']:
                    if d_col in df_export.columns:
                        df_export[d_col] = df_export[d_col].dt.strftime('%d/%m/%Y').fillna("")

                import io
                from openpyxl.styles import PatternFill, Font, Alignment

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Data Hauling')

                    # Mengambil worksheet untuk pewarnaan
                    workbook = writer.book
                    worksheet = writer.sheets['Data Hauling']

                    # Mewarnai Header (Biru gelap dengan teks putih)
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)

                    for cell in worksheet[1]:  # Baris 1 adalah header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Proteksi Performa: Hanya lakukan styling/pewarnaan zebra jika data di bawah 2000 baris
                    if len(df_export) <= 2000:
                        # Mewarnai baris selang-seling (Zebra striping - abu pucat)
                        gray_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
                        for row_idx in range(2, worksheet.max_row + 1):
                            if row_idx % 2 == 0:
                                for col_idx in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = gray_fill

                        # Menyesuaikan lebar kolom secara otomatis
                        for col in worksheet.columns:
                            max_length = 0
                            column_letter = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40) # Maksimal lebar 40

                excel_data = output.getvalue()

                # Menempatkan tombol download ke dalam placeholder di samping tombol Cari
                with download_placeholder:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    st.download_button(
                        label=":material/download: Excel",
                        data=excel_data,
                        file_name=f"Report_Hauling_{time.strftime('%Y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True
                    )

                # Ubah header menjadi kapital semua agar terlihat lebih tebal/bold
                col_cfg = {
                    "weight_gross": st.column_config.NumberColumn("WEIGHT GROSS", format="%.3f"),
                    "weight_empty": st.column_config.NumberColumn("WEIGHT EMPTY", format="%.3f"),
                    "weight_nett": st.column_config.NumberColumn("WEIGHT NETT", format="%.3f"),
                    "tonage": st.column_config.NumberColumn("TONAGE", format="%.2f"),
                    "date": st.column_config.DateColumn("DATE", format="DD/MM/YYYY"),
                    "loading_date": st.column_config.DateColumn("LOADING DATE", format="DD/MM/YYYY")
                }
                # Kapitalisasi sisa kolom lainnya
                for c in df.columns:
                    if c not in col_cfg:
                        col_cfg[c] = st.column_config.Column(c.upper().replace('_', ' '))

                st.dataframe(
                    df, 
                    use_container_width=True, 
                    hide_index=True, 
                    height=700,
                    column_config=col_cfg
                )
            else:
                st.info("Data hauling tidak ditemukan. Silakan ubah filter Anda.")

        with tab2:
            st.header("Update Data via Upload Excel")
            st.info("Pilih file `DAILY COAL HAULING MASTER REKAP.xlsx`")
            uploaded_file = st.file_uploader("Upload File Excel Hauling", type=["xlsx", "xlsb", "xls"], key="hauling")
            if uploaded_file is not None:
                if st.button(":material/cloud_upload: Import ke Server", type="primary"):
                    with st.status("Memproses Import...", expanded=True) as status:
                        st.write("Membaca file Excel (Sheet: Timbangan)...")
                        try:
                            import numpy as np
                            # Coba baca dengan asumsi header di baris ke-4 (index 3) seperti template awal
                            df_upload = pd.read_excel(uploaded_file, sheet_name='Timbangan', header=3)
                            df_upload.columns = df_upload.columns.astype(str).str.lower().str.strip().str.replace(r'\s+', '_', regex=True)

                            # Jika tidak ketemu, kemungkinan user mengupload file yang baris atasnya sudah dihapus (header di baris 1)
                            if 'voucher_number' not in df_upload.columns:
                                uploaded_file.seek(0) # Reset pointer file
                                df_upload = pd.read_excel(uploaded_file, sheet_name='Timbangan', header=0)
                                df_upload.columns = df_upload.columns.astype(str).str.lower().str.strip().str.replace(r'\s+', '_', regex=True)

                            st.write("Menjalankan Auto-Validator...")

                            # 2. Kolom database yang diharapkan
                            db_cols = ['jml', 'day', 'date', 'shift', 'loading_date', 'voucher_number', 'pit', 'block', 'seam', 'product', 'concat', 'vendor', 'unit_type', 'unit_id', 'payload_arrival_time', 'payload_embark_time', 'weight_gross', 'weight_empty', 'weight_nett', 'destination', 'route', 'loader_id', 'tonage']

                            # 3. Filter hanya kolom yang relevan & buang baris kosong
                            available_cols = [c for c in db_cols if c in df_upload.columns]
                            df_insert = df_upload[available_cols].copy()

                            if 'voucher_number' not in df_insert.columns:
                                raise ValueError("Kolom 'Voucher Number' tidak ditemukan! Pastikan header tabel berada di baris ke-4 dan namanya sesuai template.")

                            df_insert.dropna(subset=['voucher_number'], inplace=True)

                            # 4. Format Tanggal YYYY-MM-DD
                            if 'date' in df_insert.columns:
                                df_insert['date'] = pd.to_datetime(df_insert['date'], errors='coerce').dt.strftime('%Y-%m-%d')
                            if 'loading_date' in df_insert.columns:
                                df_insert['loading_date'] = pd.to_datetime(df_insert['loading_date'], errors='coerce').dt.strftime('%Y-%m-%d')

                            # 5. Parsing Time
                            for t_col in ['payload_arrival_time', 'payload_embark_time']:
                                if t_col in df_insert.columns:
                                    df_insert[t_col] = df_insert[t_col].astype(str).apply(lambda x: x.split()[-1] if x not in ['NaT', 'nan', 'None'] else None)

                            # 6. Bersihkan NaN menjadi None untuk MySQL
                            df_insert = df_insert.replace({np.nan: None, pd.NaT: None, 'nan': None})

                            st.write(f"Mempersiapkan {len(df_insert)} baris untuk diimport...")

                            # 7. Eksekusi Batch INSERT IGNORE
                            placeholders = ", ".join(["%s"] * len(available_cols))
                            cols_str = ", ".join(available_cols)
                            query_insert = f"INSERT IGNORE INTO coal_hauling ({cols_str}) VALUES ({placeholders})"

                            data_to_insert = [tuple(x) for x in df_insert.to_numpy()]

                            st.write("Mengimport batch data ke MySQL...")
                            success = database.execute_many_query(query_insert, data_to_insert)

                            if success:
                                status.update(label="Import Selesai!", state="complete", expanded=False)
                                st.success(f"{len(df_insert)} Data Hauling berhasil diproses (duplikat diabaikan)!")
                                clear_all_cache()  # Hapus semua cache agar data segar
                                time.sleep(2)
                                if "hauling" in st.session_state:
                                    del st.session_state["hauling"]
                                st.rerun()
                            else:
                                status.update(label="Import Gagal!", state="error", expanded=True)
                        except Exception as e:
                            status.update(label="Terjadi Kesalahan", state="error", expanded=True)
                            st.error(f"Error membaca/mengunggah Excel: {e}")

        with tab3:
            st.header("Rollback / Delete Batch")
            st.error("⚠️ Data yang dihapus tidak bisa dikembalikan. Gunakan hanya jika ada revisi salah ketik Excel.")

            c1, c2 = st.columns(2)
            del_date = c1.date_input("Pilih Tanggal yang Salah (Bisa Range)", value=[], key="h_date")
            del_shift = c2.selectbox("Pilih Shift", ["Semua", "DAY", "NIGHT"], key="h_shift")

            confirm_del = st.checkbox("Saya yakin ingin menghapus data secara permanen", key="h_confirm")

            if st.button(":material/delete_sweep: Hapus Batch Hauling", type="primary", disabled=not confirm_del):
                if not del_date:
                    st.warning("Silakan pilih tanggal terlebih dahulu.")
                else:
                    params = []
                    query = "DELETE FROM coal_hauling WHERE "

                    if len(del_date) == 2:
                        query += "date BETWEEN %s AND %s"
                        params.extend([del_date[0], del_date[1]])
                    elif len(del_date) == 1:
                        query += "date = %s"
                        params.append(del_date[0])

                    if del_shift != "Semua":
                        query += " AND shift = %s"
                        params.append(del_shift)

                    if database.execute_query(query, tuple(params)):
                        st.success("Data berhasil dihapus dari server! Silakan lakukan upload ulang di tab Upload/Import.")
                        clear_all_cache()  # Hapus semua cache agar data segar
                        time.sleep(2)
                        for key in ["h_date", "h_shift", "h_confirm"]:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.rerun()
                    else:
                        st.error("Terjadi kegagalan saat menghapus data.")



    elif sub_menu == "Coal Transit":
        tab1, tab2, tab3 = st.tabs([":material/monitoring: View Data", ":material/cloud_upload: Upload / Import", ":material/delete_sweep: Rollback / Delete Batch"])
        
        with tab1:
            with st.expander(":material/filter_alt: Filter & Cari Data Transit", expanded=True):
                cols = st.columns([1.5, 0.9, 1, 1, 1, 1.2, 0.7, 0.5, 0.8])
                with cols[0]:
                    f_date = st.date_input("Tanggal Transit", value=[], key="t_date")
                with cols[1]:
                    f_shift = st.selectbox("Shift", ["Semua", "DAY", "NIGHT", "Day", "Night"], key="t_shift")
                with cols[2]:
                    f_pit = st.selectbox("Pit", get_unique_transit_pits(), key="t_pit")
                with cols[3]:
                    f_digger = st.selectbox("Digger", get_unique_transit_diggers(), key="t_digger")
                with cols[4]:
                    f_unit = st.selectbox("Unit Code", get_unique_transit_unit_codes(), key="t_unit")
                with cols[5]:
                    f_search = st.text_input("Pencarian", placeholder="Ketik kata kunci...", key="t_search")
                with cols[6]:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    btn_search = st.button(":material/search: Cari", type="primary", use_container_width=True, key="t_btn")
                with cols[7]:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    if st.button(":material/refresh:", use_container_width=True, help="Tarik ulang data segar dari database", key="t_ref"):
                        clear_all_cache()
                        st.rerun()
                with cols[8]:
                    download_placeholder_transit = st.container()

            date_start = f_date[0] if len(f_date) >= 1 else None
            date_end = f_date[1] if len(f_date) >= 2 else None
            
            with st.spinner("Mengambil data..."):
                df_transit = fetch_transit_data(date_start, date_end)
            
            if not df_transit.empty:
                if f_shift != "Semua":
                    df_transit = df_transit[df_transit['shift'].str.upper() == f_shift.upper()]
                if f_pit != "Semua":
                    df_transit = df_transit[df_transit['pit'] == f_pit]
                if f_digger != "Semua":
                    df_transit = df_transit[df_transit['digger'] == f_digger]
                if f_unit != "Semua":
                    df_transit = df_transit[df_transit['unit_code'] == f_unit]
                if f_search:
                    mask = df_transit.astype(str).apply(lambda x: x.str.contains(f_search, case=False, na=False)).any(axis=1)
                    df_transit = df_transit[mask]
                
            if not df_transit.empty:
                for date_col in ['date']:
                    if date_col in df_transit.columns:
                        df_transit[date_col] = pd.to_datetime(df_transit[date_col], errors='coerce')
                
                df_export = df_transit.copy()
                for d_col in ['date']:
                    if d_col in df_export.columns:
                        df_export[d_col] = df_export[d_col].dt.strftime('%d/%m/%Y').fillna("")

                import io
                from openpyxl.styles import PatternFill, Font, Alignment
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Data Transit')
                    workbook = writer.book
                    worksheet = writer.sheets['Data Transit']
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if len(df_export) <= 2000:
                        gray_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
                        for row_idx in range(2, worksheet.max_row + 1):
                            if row_idx % 2 == 0:
                                for col_idx in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = gray_fill
                        for col in worksheet.columns:
                            max_length = 0
                            column_letter = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except: pass
                            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

                excel_data = output.getvalue()
                
                with download_placeholder_transit:
                    st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                    st.download_button(
                        label=":material/download: Excel",
                        data=excel_data,
                        file_name=f"Report_Transit_{time.strftime('%Y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        key="t_dl"
                    )
                
                col_cfg = {
                    "date": st.column_config.DateColumn("DATE", format="DD/MM/YYYY")
                }
                for c in df_transit.columns:
                    if c not in col_cfg:
                        col_cfg[c] = st.column_config.Column(c.upper().replace('_', ' '))

                st.dataframe(
                    df_transit, 
                    use_container_width=True, 
                    hide_index=True, 
                    height=700,
                    column_config=col_cfg
                )
            else:
                st.info("Data transit tidak ditemukan. Silakan ubah filter Anda.")

        with tab2:
            st.header("Update Data via Upload Excel (Transit)")
            st.info("Pilih file Excel yang mengandung sheet `Transit` atau sheet pertama.")
            uploaded_file_transit = st.file_uploader("Upload File Excel Transit", type=["xlsx", "xlsb", "xls"], key="up_transit")
            if uploaded_file_transit is not None:
                if st.button(":material/cloud_upload: Import ke Server", type="primary", key="btn_up_transit"):
                    with st.status("Memproses Import Transit...", expanded=True) as status:
                        st.write("Membaca file Excel...")
                        try:
                            import numpy as np
                            # Mencoba baca sheet 'Transit', kalau gagal baca sheet aktif
                            try:
                                df_up_t = pd.read_excel(uploaded_file_transit, sheet_name='Transit', header=None)
                            except:
                                uploaded_file_transit.seek(0)
                                df_up_t = pd.read_excel(uploaded_file_transit, header=None)
                            
                            st.write("Mencari posisi Header otomatis...")
                            # Cari baris yang memiliki kata 'unit' atau 'digger' untuk jadi header
                            header_idx = 0
                            for idx, row in df_up_t.iterrows():
                                row_str = " ".join([str(x).lower() for x in row.values])
                                if 'unit' in row_str and 'digger' in row_str:
                                    header_idx = idx
                                    break
                            
                            df_up_t.columns = df_up_t.iloc[header_idx]
                            df_up_t = df_up_t.iloc[header_idx+1:].reset_index(drop=True)
                            
                            st.write("Menjalankan Auto-Validator...")
                            df_up_t.columns = df_up_t.columns.astype(str).str.lower().str.strip().str.replace(r'\s+', '_', regex=True)
                            
                            db_cols_t = ['date', 'shift', 'unit_code', 'model', 'type', 'brand', 'user', 'seam', 'block', 'product_code', 'product_inv', 'pit', 'digger', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', 'total', 'vessel', 'netto', 'periode', 'room']
                            
                            avail_cols_t = [c for c in db_cols_t if c in df_up_t.columns]
                            if len(avail_cols_t) < 5:
                                raise ValueError("Tidak bisa menemukan kolom-kolom standar Transit di Excel ini. Pastikan format tabelnya sesuai.")
                                
                            df_ins_t = df_up_t[avail_cols_t].copy()
                            
                            # Drop baris kosong yang tidak punya date atau unit_code
                            if 'date' in df_ins_t.columns and 'unit_code' in df_ins_t.columns:
                                df_ins_t.dropna(subset=['date', 'unit_code'], how='all', inplace=True)
                            
                            if 'date' in df_ins_t.columns:
                                df_ins_t['date'] = pd.to_datetime(df_ins_t['date'], errors='coerce').dt.strftime('%Y-%m-%d')
                                
                            df_ins_t = df_ins_t.replace({np.nan: None, pd.NaT: None, 'nan': None})
                            
                            st.write(f"Mempersiapkan {len(df_ins_t)} baris Transit untuk diimport...")
                            
                            # Cek kolom yang harus dibungkus dengan backtick (seperti `1`, `2` karena angka)
                            safe_cols = [f"`{c}`" for c in avail_cols_t]
                            placeholders = ", ".join(["%s"] * len(avail_cols_t))
                            cols_str = ", ".join(safe_cols)
                            
                            query_insert = f"INSERT IGNORE INTO coal_transit ({cols_str}) VALUES ({placeholders})"
                            data_to_insert = [tuple(x) for x in df_ins_t.to_numpy()]
                            
                            st.write("Mengimport batch data ke MySQL (mencegah duplikat)...")
                            success = database.execute_many_query(query_insert, data_to_insert)
                            
                            if success:
                                status.update(label="Import Transit Selesai!", state="complete", expanded=False)
                                st.success(f"{len(df_ins_t)} Data Transit berhasil diproses (duplikat diabaikan)!")
                                clear_all_cache()
                                time.sleep(2)
                                if "up_transit" in st.session_state:
                                    del st.session_state["up_transit"]
                                st.rerun()
                            else:
                                status.update(label="Import Gagal!", state="error", expanded=True)
                        except Exception as e:
                            status.update(label="Terjadi Kesalahan", state="error", expanded=True)
                            st.error(f"Error membaca/mengunggah Excel Transit: {e}")

        with tab3:
            st.header("Rollback / Delete Batch (Transit)")
            st.error("⚠️ Data Transit yang dihapus tidak bisa dikembalikan.")
            
            c1, c2 = st.columns(2)
            del_date = c1.date_input("Pilih Tanggal yang Salah (Bisa Range)", value=[], key="td_date")
            del_shift = c2.selectbox("Pilih Shift", ["Semua", "DAY", "NIGHT", "Day", "Night"], key="td_shift")
            
            confirm_del = st.checkbox("Saya yakin ingin menghapus data Transit secara permanen", key="td_confirm")
            
            if st.button(":material/delete_sweep: Hapus Batch Transit", type="primary", disabled=not confirm_del, key="btn_del_transit"):
                if not del_date:
                    st.warning("Silakan pilih tanggal terlebih dahulu.")
                else:
                    params = []
                    q = "DELETE FROM coal_transit WHERE "
                    
                    if len(del_date) == 2:
                        q += "date BETWEEN %s AND %s"
                        params.extend([del_date[0], del_date[1]])
                    else:
                        q += "date = %s"
                        params.append(del_date[0])
                        
                    if del_shift != "Semua":
                        q += " AND shift = %s"
                        params.append(del_shift)
                        
                    with st.spinner("Menghapus data..."):
                        if database.execute_query(q, tuple(params)):
                            st.success("Batch Transit berhasil dihapus!")
                            clear_all_cache()
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error("Gagal menghapus batch Transit.")
elif menu == "Modul Overburden (OB)":
    st.title("Modul Overburden (OB)")
    st.markdown("Manajemen data pengupasan tanah (Overburden).")
    
    tab1, tab2, tab3 = st.tabs([":material/monitoring: View Data", ":material/cloud_upload: Upload Data", ":material/delete_sweep: Rollback Data"])
    
    # CSS khusus untuk sub-tab OB (warna berbeda dari tab utama)
    st.markdown("""
    <style>
        /* Sub-tab OB: Bentuk persis seperti tab utama (tanpa kotak border), 
           tapi saat diklik (aktif) warnanya menjadi Oranye. */
        div[data-testid="stTabs"] div[data-testid="stTabs"] button[data-baseweb="tab"] {
            font-size: 14px !important;
            font-weight: 600 !important;
            padding: 8px 16px !important; /* Sedikit lebih kecil agar muat 10 tab */
            border-radius: 6px 6px 0 0 !important;
            background-color: transparent !important;
            border: none !important;
            color: #31333F !important;
        }
        div[data-testid="stTabs"] div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
            background-color: #ea580c !important; /* Warna oranye saat aktif */
            color: white !important;
        }
        div[data-testid="stTabs"] div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="false"]:hover {
            background-color: #ffedd5 !important;
            color: #ea580c !important;
        }
        /* Hilangkan garis hijau (highlight) bawaan Streamlit di bawah tab aktif */
        div[data-testid="stTabs"] div[data-testid="stTabs"] div[data-baseweb="tab-highlight"] {
            display: none !important;
            background-color: transparent !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    with tab1:
        with st.expander(":material/filter_alt: Filter & Cari Data OB", expanded=True):
            cols = st.columns([1.5, 0.9, 1, 1.2, 1.2, 0.7, 0.5, 0.8])
            with cols[0]:
                f_date = st.date_input("Tanggal", value=[], key="ob_f_date")
            with cols[1]:
                f_shift = st.selectbox("Shift", get_ob_unique_shifts(), key="ob_f_shift")
            with cols[2]:
                f_pit = st.selectbox("Pit", get_ob_unique_pits(), key="ob_f_pit")
            with cols[3]:
                f_owner = st.selectbox("Owner/Subcont", get_ob_unique_owners(), key="ob_f_owner")
            with cols[4]:
                f_search = st.text_input("Pencarian", placeholder="Eqnum/Loader...", key="ob_f_search")
            with cols[5]:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                btn_search = st.button(":material/search: Cari", type="primary", use_container_width=True, key="ob_btn_search")
            with cols[6]:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                if st.button(":material/refresh:", use_container_width=True, key="ob_btn_refresh", help="Tarik ulang data segar dari database"):
                    clear_all_cache()
                    st.rerun()
            with cols[7]:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                # Tombol download akan di-render di sini nanti setelah data di-fetch
                download_placeholder = st.container()

        # Konversi filter date
        date_start = f_date[0].strftime('%Y-%m-%d') if len(f_date) >= 1 else None
        date_end = f_date[1].strftime('%Y-%m-%d') if len(f_date) >= 2 else None
        
        ob_tabs_list = [
            ("db_ob", "ob_ob"),
            ("db_ob Inpit", "ob_ob_inpit"),
            ("db_event", "ob_event"),
            ("db_Event MGE", "ob_event_mge"),
            ("db_problem prodty", "ob_problem_prodty"),
            ("Performance Subcont", "ob_performance_subcont"),
            ("Weather", "ob_weather"),
            ("Freq Weather", "ob_freq_weather"),
            ("Volume OB by JS", "ob_volume_ob_by_js"),
            ("db_Material", "ob_material")
        ]
        
        # FETCH SEMUA DATA DULU SEBELUM BIKIN TAB
        all_dfs = {}
        with st.spinner("Memuat data OB..."):
            for name, table_name in ob_tabs_list:
                sql_query = f"SELECT * FROM {table_name} WHERE 1=1"
                sql_params = []
                if date_start and date_end:
                    sql_query += " AND date BETWEEN %s AND %s"
                    sql_params.extend([date_start, date_end])
                elif date_start:
                    sql_query += " AND date = %s"
                    sql_params.append(date_start)
                
                sql_query += " ORDER BY date DESC, id DESC LIMIT 1000"
                df_view = database.fetch_data(sql_query, tuple(sql_params) if sql_params else None)
                
                if not df_view.empty:
                    if f_shift != "Semua" and 'shift' in df_view.columns:
                        df_view = df_view[df_view['shift'].astype(str) == f_shift]
                    if f_pit != "Semua" and 'pit' in df_view.columns:
                        df_view = df_view[df_view['pit'] == f_pit]
                    if f_owner != "Semua" and 'owner' in df_view.columns:
                        df_view = df_view[df_view['owner'] == f_owner]
                    
                    if f_search:
                        mask = df_view.astype(str).apply(lambda x: x.str.contains(f_search, case=False, na=False)).any(axis=1)
                        df_view = df_view[mask]
                            
                all_dfs[name] = df_view
                
        # BIKIN FILE EXCEL DAN RENDER TOMBOLNYA
        if all_dfs:
            import io
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for sheet_name, df_sheet in all_dfs.items():
                    if df_sheet.empty:
                        pd.DataFrame({'Data': []}).to_excel(writer, index=False, sheet_name=sheet_name[:31])
                    else:
                        df_sheet.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            
            with download_placeholder:
                st.download_button(
                    label="📥 Excel",
                    data=output.getvalue(),
                    file_name=f"OB_Data_{date_start or 'Semua'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="ob_dl_all"
                )

        # BARU BIKIN TAB DAN TAMPILKAN DATA
        ob_tabs = st.tabs([name for name, _ in ob_tabs_list])
        
        for i, tab in enumerate(ob_tabs):
            with tab:
                tab_name = ob_tabs_list[i][0]
                table_name = ob_tabs_list[i][1]
                df_view = all_dfs[tab_name]
                
                if not df_view.empty:
                    st.dataframe(df_view, use_container_width=True, hide_index=True)
                    st.info(f"Menampilkan {len(df_view)} baris hasil filter (Maks. 1000).")
                else:
                    st.info(f"Data di tabel {table_name} masih kosong atau tidak sesuai kriteria filter.")

    with tab2:
        st.header("Upload Data OB (10 Tabel)")
        st.markdown("Pastikan file Excel yang Anda unggah memiliki **10 sheet** dengan nama yang sama persis dengan nama tabel.")
        
        uploaded_file = st.file_uploader("Upload Excel OB", type=["xlsx"], key="ob_upload")
        
        if uploaded_file:
            if st.button("🚀 Proses Upload 10 Tabel OB", type="primary", use_container_width=True):
                with st.spinner("Membaca dan memvalidasi 10 sheet OB..."):
                    try:
                        xls = pd.ExcelFile(uploaded_file)
                        sheet_names = xls.sheet_names
                        
                        sheet_mapping = {
                            "db_ob": "ob_ob",
                            "db_ob inpit": "ob_ob_inpit",
                            "db_event": "ob_event",
                            "db_event mge": "ob_event_mge",
                            "db_problem prodty": "ob_problem_prodty",
                            "performance subcont": "ob_performance_subcont",
                            "weather": "ob_weather",
                            "freq weather": "ob_freq_weather",
                            "volume ob by js": "ob_volume_ob_by_js",
                            "db_material": "ob_material"
                        }
                        
                        success_count = 0
                        error_msgs = []
                        
                        for sheet in sheet_names:
                            sheet_lower = sheet.strip().lower()
                            if sheet_lower in sheet_mapping:
                                table = sheet_mapping[sheet_lower]
                                try:
                                    df = pd.read_excel(xls, sheet_name=sheet)
                                    if df.empty:
                                        error_msgs.append(f"Sheet '{sheet}' kosong, dilewati.")
                                        continue
                                        
                                    df.columns = [str(c).strip().lower() for c in df.columns]
                                    
                                    # Normalisasi nama kolom sesuai dengan skema database
                                    col_mapping = {
                                        'sub material': 'sub_material',
                                        'volume adjst by js': 'volume_adjst_by_js',
                                        'desc.': 'description',
                                        'start': 'start_time',
                                        'stop': 'stop_time',
                                        'desc.lama': 'desc_lama',
                                        "problem prod'ty": 'problem_prodty',
                                        "problem prod'ty.lama": 'problem_prodty_lama',
                                        'pa fmc': 'pa_fmc',
                                        'volume js': 'volume_js',
                                        'type material': 'type_material'
                                    }
                                    df.rename(columns=col_mapping, inplace=True)
                                    
                                    # Bersihkan karakter sisa jika masih ada
                                    df.columns = [c.replace(' ', '_').replace('.', '').replace("'", "") for c in df.columns]
                                    
                                    # Format datetime columns to string
                                    for col in df.select_dtypes(include=['datetime64', 'datetimetz']).columns:
                                        if col == 'date':
                                            df[col] = df[col].dt.strftime('%Y-%m-%d')
                                        else:
                                            df[col] = df[col].dt.strftime('%H:%M:%S')
                                            
                                    # Fallback manual date parse if not datetime format in excel
                                    if 'date' in df.columns and df['date'].dtype == 'O':
                                        try:
                                            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
                                        except:
                                            pass
                                            
                                    # Ubah NaN menjadi None agar bisa masuk ke MySQL sebagai NULL
                                    df = df.where(pd.notnull(df), None)
                                    
                                    # Buat tuple data
                                    data_to_insert = [tuple(x) for x in df.to_numpy()]
                                    
                                    # Generate SQL (Gunakan INSERT IGNORE agar data duplikat dilewati tanpa error)
                                    # Generate SQL (Gunakan INSERT IGNORE agar data duplikat dilewati tanpa error)
                                    # Tambahkan backtick pada nama kolom agar aman jika ada keyword bentrok
                                    columns = ", ".join([f"`{c}`" for c in df.columns])
                                    placeholders = ", ".join(["%s"] * len(df.columns))
                                    sql = f"INSERT IGNORE INTO {table} ({columns}) VALUES ({placeholders})"
                                    
                                    # Eksekusi batch
                                    if database.execute_many_query(sql, data_to_insert):
                                        success_count += 1
                                    else:
                                        error_msgs.append(f"Gagal execute batch untuk tabel {table}.")
                                except Exception as e:
                                    error_msgs.append(f"Gagal memproses sheet {sheet}: {e}")
                            else:
                                error_msgs.append(f"Sheet '{sheet}' tidak dikenali, diabaikan.")
                        
                        if success_count > 0:
                            if success_count == len(sheet_mapping):
                                st.success(f"✅ Semua {success_count} sheet OB berhasil diunggah!")
                            else:
                                st.warning(f"⚠️ Berhasil mengunggah {success_count} dari {len(sheet_mapping)} sheet.")
                            clear_all_cache()
                        else:
                            st.error("❌ Gagal mengunggah data.")
                            
                        if error_msgs:
                            with st.expander("Lihat detail error"):
                                for msg in error_msgs:
                                    st.write(f"- {msg}")
                                    
                    except Exception as e:
                        st.error(f"❌ Terjadi kesalahan sistem: {e}")

    with tab3:
        st.header("Hapus Data (Rollback) OB")
        st.warning("Menghapus data di sini akan memengaruhi ke-10 tabel OB secara bersamaan!")
        
        del_date = st.date_input("Pilih Tanggal yang ingin dihapus dari ke-10 Tabel OB:", key="ob_del_date")
        del_shift = st.selectbox("Pilih Shift (Opsional):", ["Semua Shift", "1", "2"], key="ob_del_shift")
        
        st.error(f"⚠️ Anda yakin ingin menghapus data OB tanggal {del_date} {'(Semua Shift)' if del_shift == 'Semua Shift' else f'(Shift {del_shift})'}?")
        
        confirm = st.text_input("Ketik 'HAPUS' untuk konfirmasi", key="ob_del_confirm")
        
        if st.button("Hapus Data Permanen", type="primary", disabled=(confirm != "HAPUS")):
            with st.spinner("Menghapus data dari 10 tabel OB..."):
                tables_to_delete = [
                    "ob_event", "ob_event_mge", "ob_freq_weather", "ob_ob", "ob_ob_inpit",
                    "ob_performance_subcont", "ob_problem_prodty", "ob_volume_ob_by_js", "ob_weather",
                    "ob_material"
                ]
                
                success_del = 0
                for table in tables_to_delete:
                    if del_shift == "Semua Shift":
                        sql = f"DELETE FROM {table} WHERE date = %s"
                        params = (del_date,)
                    else:
                        # Cek apakah tabel punya kolom shift
                        df_check = database.fetch_data(f"SHOW COLUMNS FROM {table} LIKE 'shift'")
                        if not df_check.empty:
                            sql = f"DELETE FROM {table} WHERE date = %s AND shift = %s"
                            params = (del_date, del_shift)
                        else:
                            # Jika tidak ada kolom shift, hanya hapus berdasarkan tanggal
                            sql = f"DELETE FROM {table} WHERE date = %s"
                            params = (del_date,)
                    
                    if database.execute_query(sql, params):
                        success_del += 1
                        
                if success_del == len(tables_to_delete):
                    st.success(f"✅ Data tanggal {del_date} berhasil dihapus dari semua tabel OB.")
                    clear_all_cache()
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(f"⚠️ Berhasil menghapus dari {success_del}/{len(tables_to_delete)} tabel.")


elif menu == "Modul MS Kontrak":
    st.title("Modul MS Kontrak")
    st.markdown("Manajemen data Mine Survey dan volume kontrak.")
    
    tab1, tab2 = st.tabs([":material/monitoring: View Survey", ":material/cloud_upload: Upload Survey"])
    
    with tab1:
        st.info("Laporan hasil survey (MS) akan tampil di sini.")
    with tab2:
        st.header("Upload Laporan Survey")
        st.file_uploader("Upload Excel MS", type=["xlsx"], key="ms")


elif menu == "Modul Fuel":
    st.title("Modul Fuel")
    st.markdown("Manajemen pemakaian dan logistik bahan bakar (Fuel).")
    
    tab1, tab2, tab3 = st.tabs([":material/monitoring: View Data", ":material/cloud_upload: Upload / Import", ":material/delete_sweep: Rollback / Delete Batch"])
    
    with tab1:
        with st.expander(":material/filter_alt: Filter & Cari Data", expanded=True):
            cols = st.columns([1.5, 0.9, 1, 1, 1, 1.2, 0.7, 0.5, 0.8])
            with cols[0]:
                f_date = st.date_input("Tanggal", value=[], key="f_date_fuel")
            with cols[1]:
                f_shift = st.selectbox("Shift", ["Semua", "DAY", "NIGHT"], key="f_shift_fuel")
            with cols[2]:
                f_loc = st.selectbox("Lokasi", get_unique_locations_fuel(), key="f_loc_fuel")
            with cols[3]:
                f_vendor = st.selectbox("Vendor", get_unique_vendors_fuel(), key="f_vendor_fuel")
            with cols[4]:
                f_unit = st.selectbox("Unit", get_unique_units_fuel(), key="f_unit_fuel")
            with cols[5]:
                f_search = st.text_input("Pencarian", placeholder="Voucher/ID...", key="f_search_fuel")
            with cols[6]:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                btn_search = st.button(":material/search: Cari", type="primary", use_container_width=True, key="btn_search_fuel")
            with cols[7]:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                if st.button(":material/refresh:", use_container_width=True, help="Tarik ulang data segar dari database", key="btn_refresh_fuel"):
                    clear_all_cache()
                    st.rerun()
            with cols[8]:
                download_placeholder = st.container()

        date_start = f_date[0] if len(f_date) >= 1 else None
        date_end = f_date[1] if len(f_date) >= 2 else None
        
        with st.spinner("Mengambil data..."):
            df = fetch_fuel_data(date_start, date_end)
        
        if not df.empty:
            if f_shift != "Semua":
                df = df[df['shift'] == f_shift]
            if f_loc != "Semua":
                df = df[df['alocation'] == f_loc]
            if f_vendor != "Semua":
                df = df[df['vendor'] == f_vendor]
            if f_unit != "Semua":
                df = df[df['unit_code'] == f_unit]
            if f_search:
                mask = df.astype(str).apply(lambda x: x.str.contains(f_search, case=False, na=False)).any(axis=1)
                df = df[mask]
            
        if not df.empty:
            for col in ['time']:
                if col in df.columns:
                    df[col] = df[col].astype(str).apply(lambda x: x.split()[-1] if x != 'NaT' and x != 'None' else '')
            
            for num_col in ['km', 'hm', 'refueling']:
                if num_col in df.columns:
                    df[num_col] = pd.to_numeric(df[num_col], errors='coerce')
                    
            for date_col in ['date']:
                if date_col in df.columns:
                    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                    
            df_export = df.copy()
            for col in ['km', 'hm', 'refueling']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{float(x):.2f}" if pd.notnull(x) else "")
            
            for d_col in ['date']:
                if d_col in df_export.columns:
                    df_export[d_col] = df_export[d_col].dt.strftime('%d/%m/%Y').fillna("")

            import io
            from openpyxl.styles import PatternFill, Font, Alignment
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Data Fuel')
                
                workbook = writer.book
                worksheet = writer.sheets['Data Fuel']
                
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                if len(df_export) <= 2000:
                    gray_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
                    for row_idx in range(2, worksheet.max_row + 1):
                        if row_idx % 2 == 0:
                            for col_idx in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = gray_fill
                        
                    for col in worksheet.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

            excel_data = output.getvalue()
            
            with download_placeholder:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                st.download_button(
                    label=":material/download: Excel",
                    data=excel_data,
                    file_name=f"Report_Fuel_{time.strftime('%Y%m%d')}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                    key="dl_fuel"
                )
            
            col_cfg = {
                "km": st.column_config.NumberColumn("KM", format="%.2f"),
                "hm": st.column_config.NumberColumn("HM", format="%.2f"),
                "refueling": st.column_config.NumberColumn("REFUELING", format="%.2f"),
                "date": st.column_config.DateColumn("DATE", format="DD/MM/YYYY"),
                "id": None
            }
            for c in df.columns:
                if c not in col_cfg:
                    col_cfg[c] = st.column_config.Column(c.upper().replace('_', ' '))

            st.dataframe(
                df, 
                use_container_width=True, 
                hide_index=True, 
                height=700,
                column_config=col_cfg
            )
        else:
            st.info("Data fuel tidak ditemukan. Silakan ubah filter Anda.")

    with tab2:
        st.header("Update Data via Upload Excel")
        st.info("Pilih file Excel pemakaian Fuel")
        uploaded_file = st.file_uploader("Upload File Excel Fuel", type=["xlsx", "xlsb", "xls"], key="fuel_upload")
        if uploaded_file is not None:
            if st.button(":material/cloud_upload: Import ke Server", type="primary", key="btn_import_fuel"):
                with st.status("Memproses Import...", expanded=True) as status:
                    st.write("Membaca file Excel...")
                    try:
                        import numpy as np
                        df_upload = pd.read_excel(uploaded_file, sheet_name='fuel', header=0)
                        
                        st.write("Menjalankan Auto-Validator...")
                        df_upload.columns = df_upload.columns.astype(str).str.lower().str.strip().str.replace(r'\s+', '_', regex=True)
                        
                        db_cols = ['unit_fix', 'date', 'periode', 'shift', 'time', 'reg_no', 'unit_code', 'unit_model', 'unit_type', 'brand', 'vendor', 'alocation', 'km', 'hm', 'fm_awal', 'fm_akhir', 'refueling', 'source', 'location', 'operator', 'fuelman', 'no_voucher']
                        
                        available_cols = [c for c in db_cols if c in df_upload.columns]
                        df_insert = df_upload[available_cols].copy()
                        if 'no_voucher' in df_insert.columns:
                            df_insert.dropna(subset=['no_voucher'], inplace=True)
                        elif 'unit_code' in df_insert.columns:
                            df_insert.dropna(subset=['unit_code'], inplace=True)
                            
                        if 'date' in df_insert.columns:
                            df_insert['date'] = pd.to_datetime(df_insert['date'], errors='coerce').dt.strftime('%Y-%m-%d')
                            
                        if 'time' in df_insert.columns:
                            df_insert['time'] = df_insert['time'].astype(str).apply(lambda x: x.split()[-1] if x not in ['NaT', 'nan', 'None'] else None)
                        
                        df_insert = df_insert.replace({np.nan: None, pd.NaT: None, 'nan': None})
                        
                        st.write(f"Mempersiapkan {len(df_insert)} baris untuk diimport...")
                        
                        if len(df_insert) > 0:
                            placeholders = ", ".join(["%s"] * len(available_cols))
                            cols_str = ", ".join(available_cols)
                            query_insert = f"INSERT IGNORE INTO fuel ({cols_str}) VALUES ({placeholders})"
                            
                            data_to_insert = [tuple(x) for x in df_insert.to_numpy()]
                            
                            st.write("Mengimport batch data ke MySQL...")
                            success = database.execute_many_query(query_insert, data_to_insert)
                            
                            if success:
                                status.update(label="Import Selesai!", state="complete", expanded=False)
                                st.success(f"{len(df_insert)} Data Fuel berhasil diproses (duplikat diabaikan)!")
                                clear_all_cache() 
                                time.sleep(2)
                                if "fuel_upload" in st.session_state:
                                    del st.session_state["fuel_upload"]
                                st.rerun()
                            else:
                                status.update(label="Import Gagal!", state="error", expanded=True)
                        else:
                            status.update(label="Import Gagal!", state="error", expanded=True)
                            st.error("Tidak ada data valid yang ditemukan.")
                    except Exception as e:
                        status.update(label="Terjadi Kesalahan", state="error", expanded=True)
                        st.error(f"Error membaca/mengunggah Excel: {e}")

    with tab3:
        st.header("Rollback / Delete Batch")
        st.error("⚠️ Data yang dihapus tidak bisa dikembalikan. Gunakan hanya jika ada revisi salah ketik Excel.")
        
        c1, c2 = st.columns(2)
        del_date = c1.date_input("Pilih Tanggal yang Salah (Bisa Range)", value=[], key="f_del_date")
        del_shift = c2.selectbox("Pilih Shift", ["Semua", "DAY", "NIGHT"], key="f_del_shift")
        
        confirm_del = st.checkbox("Saya yakin ingin menghapus data secara permanen", key="f_confirm_del")
        
        if st.button(":material/delete_sweep: Hapus Batch Fuel", type="primary", disabled=not confirm_del, key="btn_del_fuel"):
            if not del_date:
                st.warning("Silakan pilih tanggal terlebih dahulu.")
            else:
                params = []
                query = "DELETE FROM fuel WHERE "
                
                if len(del_date) == 2:
                    query += "date BETWEEN %s AND %s"
                    params.extend([del_date[0], del_date[1]])
                elif len(del_date) == 1:
                    query += "date = %s"
                    params.append(del_date[0])
                    
                if del_shift != "Semua":
                    query += " AND shift = %s"
                    params.append(del_shift)
                    
                if database.execute_query(query, tuple(params)):
                    st.success("Data berhasil dihapus dari server! Silakan lakukan upload ulang di tab Upload/Import.")
                    clear_all_cache() 
                    time.sleep(2)
                    for key in ["f_del_date", "f_del_shift", "f_confirm_del"]:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()
                else:
                    st.error("Terjadi kegagalan saat menghapus data.")
